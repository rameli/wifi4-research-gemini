import openpyxl
from openpyxl.styles import Font, Alignment
import re
import os

def parse_accumulator(file_path):
    with open(file_path, 'r') as f:
        content = f.read()
    
    # Split by parameter entries
    entries = re.split(r'\n### ', content)
    parameters = []
    
    for entry in entries:
        if not entry.strip() or entry.startswith('#'):
            continue
            
        lines = entry.strip().split('\n')
        name = lines[0].strip()
        param_data = {"Parameter Name": name}
        
        for line in lines[1:]:
            match = re.match(r'- \*\*([^*]+)\*\*: (.*)', line)
            if match:
                key = match.group(1).strip()
                value = match.group(2).strip()
                param_data[key] = value
        
        parameters.append(param_data)
    
    return parameters

def create_xlsx():
    accumulator_path = "/home/rameli/work/wifi4-research-gemini/log/phase3-accumulator.md"
    if not os.path.exists(accumulator_path):
        print(f"Error: {accumulator_path} not found.")
        return

    parameters = parse_accumulator(accumulator_path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phase 3 Verified Parameters"

    # Define headers based on the schema
    headers = [
        "Parameter Name",
        "Official IEEE name",
        "Official IEEE description",
        "Section",
        "Page(s)",
        "Table/Figure",
        "HT context confirmed",
        "802.11n confirmed",
        "Criticality verified",
        "Conflicts resolved",
        "Verification sources",
        "Notes"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row, param in enumerate(parameters, 2):
        for col, header in enumerate(headers, 1):
            val = param.get(header, "N/A")
            ws.cell(row=row, column=col, value=val)

    output_path = "/home/rameli/work/wifi4-research-gemini/phase3-verified-parameters.xlsx"
    wb.save(output_path)
    print(f"Successfully created {output_path}")

if __name__ == "__main__":
    create_xlsx()
