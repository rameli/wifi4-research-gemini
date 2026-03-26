import openpyxl
from openpyxl.styles import Font, Alignment

def create_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phase 2 Parameters"

    headers = [
        "Parameter Name", "Source agents", "Layer", "Criticality", 
        "Default value", "Long-distance value", "Description", 
        "Usage context", "IEEE section/page", "Has Phase 1B page ref", "Conflicts"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    parameters = [
        {
            "Parameter Name": "aSlotTime",
            "Source agents": "P-Deep-1, W1, W2, W3, K1",
            "Layer": "PHY / MAC",
            "Criticality": "Mandatory",
            "Default value": "9 µs (Short Slot), 20 µs (Long Slot)",
            "Long-distance value": "9 µs + aAirPropagationTime",
            "Description": "The time unit used by the MAC for backoff and channel access.",
            "Usage context": "Used in DCF and EDCA for backoff counters.",
            "IEEE section/page": "10.3.2.16, 10.3.7, 17.4.4",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aSIFSTime",
            "Source agents": "P-Deep-1, W1, W3, K1",
            "Layer": "PHY / MAC",
            "Criticality": "Mandatory",
            "Default value": "16 µs (5 GHz / HT), 10 µs (2.4 GHz)",
            "Long-distance value": "Fixed per PHY (10/16 µs)",
            "Description": "Short Interframe Space; the shortest interval between frame transmissions.",
            "Usage context": "Used between a data frame and its acknowledgment (ACK).",
            "IEEE section/page": "10.3.2.3.3, 17.4.4",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aDIFSTime",
            "Source agents": "P-Deep-1, W1, K1",
            "Layer": "MAC",
            "Criticality": "Mandatory",
            "Default value": "34 µs (5 GHz / HT Short Slot), 28 µs (2.4 GHz Short Slot)",
            "Long-distance value": "aSIFSTime + 2 * aSlotTime",
            "Description": "DCF Interframe Space; the minimum time the medium must be idle before a STA can transmit under DCF.",
            "Usage context": "Used for regular data and management frame transmissions.",
            "IEEE section/page": "10.3.7",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aEIFSTime",
            "Source agents": "P-Deep-1, W1, K1",
            "Layer": "MAC",
            "Criticality": "Mandatory",
            "Default value": "~94 µs (typical for 20 MHz OFDM)",
            "Long-distance value": "aSIFSTime + AckTxTime + DIFS",
            "Description": "Extended Interframe Space; used after the reception of a corrupted frame.",
            "Usage context": "Prevents a STA from interfering with an ACK that might be sent by another STA.",
            "IEEE section/page": "10.3.7",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11GuardInterval",
            "Source agents": "P-Deep-1, W1, W3, W5, K1, K3",
            "Layer": "PHY",
            "Criticality": "Mandatory",
            "Default value": "800 ns (Long GI)",
            "Long-distance value": "800 ns (Long GI preferred for stability)",
            "Description": "The circular prefix prepended to an OFDM symbol to mitigate Inter-Symbol Interference (ISI).",
            "Usage context": "Used in all OFDM/HT transmissions by default.",
            "IEEE section/page": "17.3.2.4, 19.3.6",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aSymbolTime",
            "Source agents": "P-Deep-1, W1",
            "Layer": "PHY",
            "Criticality": "Mandatory",
            "Default value": "4 µs (Long GI), 3.6 µs (Short GI)",
            "Long-distance value": "4 µs (Long GI preferred)",
            "Description": "Total duration of one OFDM symbol including the guard interval.",
            "Usage context": "Fundamental unit of data transmission in OFDM/HT.",
            "IEEE section/page": "17.3.2.4, 19.3.6",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aSignalExtension",
            "Source agents": "P-Deep-1",
            "Layer": "PHY",
            "Criticality": "Mandatory",
            "Default value": "6 µs",
            "Long-distance value": "6 µs",
            "Description": "A period of no transmission added after the end of the last symbol of an OFDM frame.",
            "Usage context": "Used in ERP-OFDM and HT formats for legacy compatibility.",
            "IEEE section/page": "10.3.8, 19.3.2",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11AckTimeout",
            "Source agents": "P-Deep-2, W1, W2, W3, W5, K1",
            "Layer": "MAC",
            "Criticality": "Mandatory",
            "Default value": "aSIFSTime + aSlotTime + aRxPHYStartDelay",
            "Long-distance value": "(2 * Propagation Delay) + SIFS + AckTxTime + Buffer",
            "Description": "The interval a STA waits for an ACK frame after transmitting an MPDU.",
            "Usage context": "Frame acknowledgment procedure.",
            "IEEE section/page": "10.3.2.11",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11CTSTimeout",
            "Source agents": "P-Deep-2, W2, K1",
            "Layer": "MAC",
            "Criticality": "Mandatory",
            "Default value": "aSIFSTime + aSlotTime + aRxPHYStartDelay",
            "Long-distance value": "(2 * Propagation Delay) + SIFS + CTSTxTime + Buffer",
            "Description": "The interval a STA waits for a CTS frame after transmitting an RTS frame.",
            "Usage context": "RTS/CTS handshake for collision avoidance.",
            "IEEE section/page": "10.3.2.9",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11RTSThreshold",
            "Source agents": "P-Deep-2, W2",
            "Layer": "MAC",
            "Criticality": "Optimization",
            "Default value": "2347",
            "Long-distance value": "Lower values (e.g., 512-1024)",
            "Description": "Determines whether an RTS/CTS exchange precedes frame transmission.",
            "Usage context": "Collision avoidance; hidden node problem mitigation.",
            "IEEE section/page": "Annex C",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11FragmentationThreshold",
            "Source agents": "P-Deep-2, W2, W3",
            "Layer": "MAC",
            "Criticality": "Optimization",
            "Default value": "2346",
            "Long-distance value": "Lower values (e.g., 512-1500)",
            "Description": "Maximum size of MPDU before fragmentation occurs.",
            "Usage context": "Frame partitioning to improve reliability in noisy environments.",
            "IEEE section/page": "Annex C",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11CoverageClass",
            "Source agents": "P-Deep-2, W2, K1",
            "Layer": "MAC Management",
            "Criticality": "Mandatory",
            "Default value": "0",
            "Long-distance value": "Integer = Ceil(Distance_in_km / 0.450)",
            "Description": "Characterizes the BSS radius and adjusts timing parameters.",
            "Usage context": "MAC timing adjustment for outdoor deployments.",
            "IEEE section/page": "Annex C",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11BeaconPeriod",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "100 TU (102.4 ms)",
            "Long-distance value": "200-1000 TU",
            "Description": "Time between successive beacon transmissions.",
            "Usage context": "BSS synchronization and discovery.",
            "IEEE section/page": "9.4.1.3",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11DTIMPeriod",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "1 to 3 beacon intervals",
            "Long-distance value": "1",
            "Description": "Number of beacon intervals between successive DTIMs.",
            "Usage context": "Broadcast/multicast traffic delivery for power-saving STAs.",
            "IEEE section/page": "9.4.2.14",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11ProbeDelay",
            "Source agents": "P-Deep-3",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "TBD",
            "Long-distance value": "TBD",
            "Description": "Time a STA waits before transmitting a Probe Request after channel switch.",
            "Usage context": "Active scanning.",
            "IEEE section/page": "Annex C",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTAMSDULimit",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "Aggregation",
            "Criticality": "Optimization",
            "Default value": "3839 or 7935 octets",
            "Long-distance value": "Lower values or Disabled",
            "Description": "Maximum size for Aggregate MAC Service Data Unit.",
            "Usage context": "Throughput optimization.",
            "IEEE section/page": "9.3.2.2",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTAMPDUThreshold",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "Aggregation",
            "Criticality": "Optimization",
            "Default value": "65,535 octets",
            "Long-distance value": "Optimized for link stability",
            "Description": "Maximum size for Aggregate MAC Protocol Data Unit.",
            "Usage context": "Throughput optimization via subframe aggregation.",
            "IEEE section/page": "9.3.2",
            "Has Phase 1B page ref": "YES",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTBlockAckWindowSize",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "Aggregation",
            "Criticality": "Optimization",
            "Default value": "64 MPDUs",
            "Long-distance value": "64 (Maximize)",
            "Description": "Number of frames that can be sent before an acknowledgment is required.",
            "Usage context": "High-throughput transmission with selective ACK.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11ADDBATimeout",
            "Source agents": "P-Deep-3, W4, K2",
            "Layer": "Aggregation",
            "Criticality": "Optimization",
            "Default value": "Negotiated",
            "Long-distance value": "Increased or Disabled (0)",
            "Description": "Duration of inactivity before a Block ACK agreement is terminated.",
            "Usage context": "Block ACK session management.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11NAVLimit",
            "Source agents": "W2",
            "Layer": "MAC",
            "Criticality": "TBD",
            "Default value": "32,767 µs",
            "Long-distance value": "Calculated",
            "Description": "Virtual carrier-sensing mechanism timer.",
            "Usage context": "Prevents other STAs from transmitting.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTMCSIndex",
            "Source agents": "W5, K3",
            "Layer": "MIMO-MCS",
            "Criticality": "Mandatory",
            "Default value": "Auto",
            "Long-distance value": "Fixed lower MCS (e.g., MCS 0-11)",
            "Description": "Index defining modulation, coding rate, and spatial streams for HT.",
            "Usage context": "Defines PHY data rate.",
            "IEEE section/page": "1.2.840.10036.4.17 (MIB)",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTNss",
            "Source agents": "W5, K3",
            "Layer": "MIMO-MCS",
            "Criticality": "Optimization",
            "Default value": "Max supported by hardware",
            "Long-distance value": "1 or 2 (Reduced to 1 for extreme range)",
            "Description": "The number of independent data streams sent simultaneously using MIMO.",
            "Usage context": "Multiplexing gain vs diversity gain.",
            "IEEE section/page": "1.2.840.10036.4.15.1.6 (MIB)",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11MaxTransmitPower",
            "Source agents": "W5, K3",
            "Layer": "Regulatory",
            "Criticality": "Mandatory",
            "Default value": "20–30 dBm",
            "Long-distance value": "Max legal/hardware (e.g., 30 dBm)",
            "Description": "Maximum power level the radio is allowed to transmit.",
            "Usage context": "Regulatory compliance and link budget.",
            "IEEE section/page": "1.2.840.10036.4.3.1.2 (MIB)",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11EIRPLimit",
            "Source agents": "W5, K3",
            "Layer": "Regulatory",
            "Criticality": "Mandatory",
            "Default value": "36 dBm",
            "Long-distance value": "PtP max",
            "Description": "Equivalent Isotropically Radiated Power.",
            "Usage context": "Regulatory limit for total radiated power.",
            "IEEE section/page": "N/A",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11AntennaGain",
            "Source agents": "W5, K3",
            "Layer": "Regulatory",
            "Criticality": "Optimization",
            "Default value": "2–5 dBi",
            "Long-distance value": "20–34 dBi",
            "Description": "Directional gain of the antenna.",
            "Usage context": "Overcoming Free Space Path Loss.",
            "IEEE section/page": "N/A",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTSTBC",
            "Source agents": "W5, K3",
            "Layer": "MIMO-MCS",
            "Criticality": "Optimization",
            "Default value": "Disabled or Auto",
            "Long-distance value": "Enabled",
            "Description": "Technique to transmit multiple copies of a data stream across several antennas.",
            "Usage context": "Improving reliability when multiplexing gain is low.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aPLCPPreambleDuration",
            "Source agents": "W1, K1",
            "Layer": "PHY",
            "Criticality": "Mandatory",
            "Default value": "36 µs (HT-Mixed)",
            "Long-distance value": "Default (remains constant)",
            "Description": "Initial part of a frame for synchronization.",
            "Usage context": "Synchronization and training.",
            "IEEE section/page": "Section 19.3.3",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11AssociationResponseTimeOut",
            "Source agents": "W4",
            "Layer": "MAC Management",
            "Criticality": "Mandatory",
            "Default value": "512 TU",
            "Long-distance value": "1000-2048 TU",
            "Description": "Time a STA waits for an Association Response.",
            "Usage context": "Initial connection handshake.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11AuthenticationResponseTimeOut",
            "Source agents": "W4",
            "Layer": "MAC Management",
            "Criticality": "Mandatory",
            "Default value": "512 TU",
            "Long-distance value": "1000 TU",
            "Description": "Time a STA waits for the next frame in authentication sequence.",
            "Usage context": "Initial 802.11 authentication phase.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11MinChannelTime",
            "Source agents": "W4",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "10-20 ms",
            "Long-distance value": "20-40 ms",
            "Description": "Min time spent on channel during active scan.",
            "Usage context": "Active scanning.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11MaxChannelTime",
            "Source agents": "W4",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "30-60 ms",
            "Long-distance value": "60-100 ms",
            "Description": "Max time spent on channel during active scan.",
            "Usage context": "Active scanning.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11HTChannelWidth",
            "Source agents": "W3, W5, K3",
            "Layer": "PHY",
            "Criticality": "Optimization / Mandatory (for range)",
            "Default value": "20 MHz",
            "Long-distance value": "20 MHz (preferred for range)",
            "Description": "Spectral bandwidth used for transmission.",
            "Usage context": "802.11n bonds two 20 MHz channels for 40 MHz.",
            "IEEE section/page": "1.2.840.10036.4.15 (MIB)",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "aPIFSTime",
            "Source agents": "K1",
            "Layer": "MAC",
            "Criticality": "Optimization",
            "Default value": "SIFS + 1 * SlotTime",
            "Long-distance value": "SIFS + (Adjusted SlotTime)",
            "Description": "Intermediate inter-frame space used by AP for priority access.",
            "Usage context": "Used in PCF or for Beacon transmission.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11ATIMWindow",
            "Source agents": "K2",
            "Layer": "MAC Management",
            "Criticality": "Optimization (IBSS only)",
            "Default value": "0 TUs",
            "Long-distance value": "5-10 TUs",
            "Description": "Announcement Traffic Indication Message window.",
            "Usage context": "IBSS power management window.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        },
        {
            "Parameter Name": "dot11ProbeResponseTimeout",
            "Source agents": "K2",
            "Layer": "MAC Management",
            "Criticality": "Optimization",
            "Default value": "10-30 ms",
            "Long-distance value": "50-100 ms",
            "Description": "Duration a STA waits for response after probe request.",
            "Usage context": "Active scanning discovery.",
            "IEEE section/page": "TBD",
            "Has Phase 1B page ref": "NO",
            "Conflicts": "None"
        }
    ]

    for row, param in enumerate(parameters, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=param[header])

    wb.save("/home/rameli/work/wifi4-research-gemini/phase2-candidate-parameters.xlsx")

if __name__ == "__main__":
    create_xlsx()
