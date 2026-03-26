import openpyxl
from openpyxl.styles import Font, Alignment
import re
import os

def parse_accumulator(file_path):
    if not os.path.exists(file_path):
        return []
        
    with open(file_path, 'r') as f:
        content = f.read()
    
    # Split by parameter entries
    entries = re.split(r'\n### ', content)
    parameters = []
    
    for entry in entries:
        if not entry.strip() or entry.startswith('#'):
            continue
            
        lines = entry.strip().split('\n')
        name = lines[0].strip()
        param_data = {"Parameter Name": name}
        
        for line in lines[1:]:
            # Match "- **Key**: Value"
            match = re.match(r'- \*\*([^*]+)\*\*: (.*)', line)
            if match:
                key = match.group(1).strip()
                value = match.group(2).strip()
                param_data[key] = value
        
        parameters.append(param_data)
    
    return parameters

def generate_final_xlsx():
    accumulator_path = "/home/rameli/work/wifi4-research-gemini/log/phase3-accumulator.md"
    parameters = parse_accumulator(accumulator_path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "802.11n Verified Parameters"

    # Define headers
    headers = [
        "Parameter Name",
        "Official IEEE name",
        "Official IEEE description",
        "Section",
        "Page(s)",
        "Table/Figure",
        "HT context confirmed",
        "802.11n confirmed",
        "Criticality verified",
        "Conflicts resolved",
        "Verification sources",
        "Notes"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Write data
    for row, param in enumerate(parameters, 2):
        for col, header in enumerate(headers, 1):
            val = param.get(header, "N/A")
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path = "/home/rameli/work/wifi4-research-gemini/parameters.xlsx"
    wb.save(output_path)
    print(f"Successfully generated {output_path}")

if __name__ == "__main__":
    generate_final_xlsx()
